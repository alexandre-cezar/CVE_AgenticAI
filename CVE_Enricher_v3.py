import os
import time
import json
import requests
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

# --- Configuration ---
KEV_URL = "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
NVD_API_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0?cveId={cve_id}"
SPREADSHEET_NAME = "Enriched CVEs.xlsx"
TIMESTAMP_FORMAT = "%Y-%m-%d %H:%M:%S"

# --- NVD API Key (replace with your key or set as an environment variable) ---
NVD_API_KEY = os.getenv('NVD_API_KEY', "YOUR_API_KEY_GOES_HERE")

# --- Throttling Configuration ---
INITIAL_WORKERS = 10

# --- Color Mapping ---
CRITICAL_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
LOW_FILL = PatternFill(start_color="008000", end_color="008000", fill_type="solid")

class RateLimitException(Exception):
    pass

def get_color_for_score(score):
    if score is None:
        return None
    if 9.0 <= score <= 10.0:
        return CRITICAL_FILL
    elif 7.0 <= score <= 8.9:
        return HIGH_FILL
    elif 4.0 <= score <= 6.9:
        return MEDIUM_FILL
    elif 0.1 <= score <= 3.9:
        return LOW_FILL
    return None

def get_color_for_severity(severity):
    if severity is None:
        return None
    severity = severity.upper()
    if severity == "CRITICAL":
        return CRITICAL_FILL
    elif severity == "HIGH":
        return HIGH_FILL
    elif severity == "MEDIUM":
        return MEDIUM_FILL
    elif severity == "LOW":
        return LOW_FILL
    return None

def fetch_nvd_data(cve_id):
    headers = {
        'User-Agent': 'Mozilla/5.0',
        'Cache-Control': 'no-cache',
        'Pragma': 'no-cache'
    }
    if NVD_API_KEY and NVD_API_KEY != "YOUR_API_KEY_GOES_HERE":
        headers['apiKey'] = NVD_API_KEY

    try:
        response = requests.get(NVD_API_URL.format(cve_id=cve_id), headers=headers, timeout=20)
        if response.status_code == 429:
            raise RateLimitException(f"Rate limit hit for {cve_id}")
        if response.status_code == 404:
            return cve_id, None, None  # Not found
        response.raise_for_status()
        data = response.json()
        if 'vulnerabilities' in data and data['vulnerabilities']:
            cve_data = data['vulnerabilities'][0]['cve']
            if 'metrics' in cve_data and 'cvssMetricV31' in cve_data['metrics']:
                cvss_data = cve_data['metrics']['cvssMetricV31'][0]['cvssData']
                return cve_id, cvss_data.get('baseScore'), cvss_data.get('baseSeverity')
        return cve_id, None, None
    except requests.exceptions.RequestException as e:
        print(f"Warning: Request for {cve_id} failed: {e}")
        return cve_id, None, None

def main():
    print("Starting the CVE enrichment process...")
    existing_df = None
    if os.path.exists(SPREADSHEET_NAME):
        print(f"Found existing spreadsheet: '{SPREADSHEET_NAME}'. It will be updated.")
        try:
            existing_df = pd.read_excel(SPREADSHEET_NAME)
        except Exception as e:
            print(f"Error reading existing spreadsheet: {e}. A new one will be created.")

    print(f"Downloading KEV list from {KEV_URL}...")
    try:
        response = requests.get(KEV_URL)
        response.raise_for_status()
        kev_data = response.json()
        print("KEV list downloaded successfully.")
    except requests.exceptions.RequestException as e:
        print(f"Error downloading KEV list: {e}")
        return

    kev_vulnerabilities = kev_data.get("vulnerabilities", [])
    if existing_df is not None:
        existing_cves = set(existing_df['CVE ID'])
        cves_to_process = [vuln for vuln in kev_vulnerabilities if vuln.get("cveID") not in existing_cves]
        if not cves_to_process:
            print("No new CVEs found in the KEV list. Exiting.")
            return
        print(f"Found {len(cves_to_process)} new CVEs to add.")
    else:
        cves_to_process = kev_vulnerabilities
        print(f"Processing {len(cves_to_process)} CVEs from the KEV list.")

    new_cve_rows = {
        vuln.get('cveID'): {
            'CVE ID': vuln.get('cveID'),
            'Vulnerability Name': vuln.get('vulnerabilityName'),
            'CVSS Score': None,
            'CVSS Severity': None,
            'EPSS Score': None,
            'Vendor Project': vuln.get('vendorProject'),
            'Vendor Product': vuln.get('product'),
            'Date Added': vuln.get('dateAdded'),
            'Due Date': vuln.get('dueDate'),
            'NVD Status': 'Unknown'
        } for vuln in cves_to_process
    }

    print("Fetching CVSS data from NVD...")
    cve_ids_to_fetch = list(new_cve_rows.keys())
    current_workers = INITIAL_WORKERS

    with tqdm(total=len(cve_ids_to_fetch), desc="Fetching NVD Data") as pbar:
        while cve_ids_to_fetch:
            rate_limit_hit = False
            next_round_to_fetch = []
            with ThreadPoolExecutor(max_workers=current_workers) as executor:
                future_to_cve = {executor.submit(fetch_nvd_data, cve_id): cve_id for cve_id in cve_ids_to_fetch}
                for future in as_completed(future_to_cve):
                    try:
                        cve_id, score, severity = future.result()
                        if score is not None:
                            new_cve_rows[cve_id]['CVSS Score'] = score
                            new_cve_rows[cve_id]['CVSS Severity'] = severity
                            new_cve_rows[cve_id]['NVD Status'] = 'Published'
                        else:
                            new_cve_rows[cve_id]['NVD Status'] = 'Not Found'
                            print(f"Note: NVD entry not found for {cve_id} (likely unpublished)")
                        pbar.update(1)
                    except RateLimitException:
                        rate_limit_hit = True
                        next_round_to_fetch.append(future_to_cve[future])
                    except Exception as e:
                        cve_id = future_to_cve[future]
                        print(f"Error for {cve_id}: {e}")
                        pbar.update(1)

            if rate_limit_hit:
                cve_ids_to_fetch = next_round_to_fetch
                pbar.set_description(f"Rate limit hit! Retrying {len(cve_ids_to_fetch)} CVEs")
                new_worker_count = max(1, int(current_workers * 0.8))
                if new_worker_count == current_workers and current_workers == 1:
                    print(f"At 1 worker and still rate-limited. Waiting 30s before retry.")
                    time.sleep(30)
                else:
                    print(f"Reducing workers from {current_workers} to {new_worker_count}. Waiting 10s.")
                    time.sleep(10)
                current_workers = new_worker_count
            else:
                cve_ids_to_fetch = []

    new_df = pd.DataFrame(list(new_cve_rows.values()))
    if existing_df is not None:
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
    else:
        combined_df = new_df

    print("\nCreating and formatting the spreadsheet...")
    with pd.ExcelWriter(SPREADSHEET_NAME, engine='openpyxl') as writer:
        combined_df.to_excel(writer, index=False, sheet_name='Enriched CVEs')
        worksheet = writer.sheets['Enriched CVEs']
        workbook = worksheet.parent

        header_map = {col: i + 1 for i, col in enumerate(combined_df.columns)}
        score_col_idx = header_map.get('CVSS Score')
        severity_col_idx = header_map.get('CVSS Severity')

        for row_index, row in enumerate(combined_df.itertuples(), 2):
            if score_col_idx:
                score_color = get_color_for_score(getattr(row, 'CVSS_Score', None))
                if score_color:
                    worksheet.cell(row=row_index, column=score_col_idx).fill = score_color
            if severity_col_idx:
                severity_color = get_color_for_severity(getattr(row, 'CVSS_Severity', None))
                if severity_color:
                    worksheet.cell(row=row_index, column=severity_col_idx).fill = severity_color

        timestamp_sheet = workbook.create_sheet("Metadata")
        timestamp_sheet['A1'] = "Last Updated"
        timestamp_sheet['B1'] = datetime.now().strftime(TIMESTAMP_FORMAT)

    print(f"Successfully created/updated '{SPREADSHEET_NAME}'")

if __name__ == "__main__":
    main()