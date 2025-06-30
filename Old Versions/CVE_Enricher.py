import os
import time
import json
import requests
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

# --- Configuration ---
KEV_URL = "https://www.cisa.gov/sites/default/files/feeds/known_exploited_vulnerabilities.json"
NVD_API_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0?cveId={cve_id}"
SPREADSHEET_NAME = "Enriched CVEs.xlsx"
TIMESTAMP_FORMAT = "%Y-%m-%d %H:%M:%S"

# --- CVE Color Mapping ---
CRITICAL_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark Red
HIGH_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
MEDIUM_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
LOW_FILL = PatternFill(start_color="008000", end_color="008000", fill_type="solid")  # Green

def get_color_for_score(score):
    """Returns the color fill for a given CVSS score."""
    if score is None:
        return None
    if 9.0 <= score <= 10.0:
        return CRITICAL_FILL
    elif 7.0 <= score <= 8.9:
        return HIGH_FILL
    elif 4.0 <= score <= 6.9:
        return MEDIUM_FILL
    elif 0.1 <= score <= 3.9:
        return LOW_FILL
    return None

def get_color_for_severity(severity):
    """Returns the color fill for a given CVSS severity."""
    if severity is None:
        return None
    severity = severity.upper()
    if severity == "CRITICAL":
        return CRITICAL_FILL
    elif severity == "HIGH":
        return HIGH_FILL
    elif severity == "MEDIUM":
        return MEDIUM_FILL
    elif severity == "LOW":
        return LOW_FILL
    return None

def fetch_nvd_data(cve_id, retries=3, delay=5):
    """Fetches CVSS score and severity from the NVD API for a given CVE ID."""
    for attempt in range(retries):
        try:
            response = requests.get(NVD_API_URL.format(cve_id=cve_id), timeout=10)
            response.raise_for_status()
            data = response.json()
            if 'vulnerabilities' in data and data['vulnerabilities']:
                cve_data = data['vulnerabilities'][0]['cve']
                if 'metrics' in cve_data and 'cvssMetricV31' in cve_data['metrics']:
                    cvss_data = cve_data['metrics']['cvssMetricV31'][0]['cvssData']
                    return cvss_data.get('baseScore'), cvss_data.get('baseSeverity')
            return None, None
        except requests.exceptions.RequestException as e:
            print(f"Warning: Request for {cve_id} failed (attempt {attempt + 1}/{retries}): {e}")
            time.sleep(delay)
    return None, None

def main():
    """Main function to perform CVE enrichment."""
    print("Starting the CVE enrichment process...")

    # --- 1. Check for existing spreadsheet ---
    existing_df = None
    if os.path.exists(SPREADSHEET_NAME):
        print(f"Found existing spreadsheet: '{SPREADSHEET_NAME}'. It will be updated.")
        try:
            existing_df = pd.read_excel(SPREADSHEET_NAME)
        except Exception as e:
            print(f"Error reading existing spreadsheet: {e}. A new one will be created.")

    # --- 2. Download KEV list ---
    print(f"Downloading KEV list from {KEV_URL}...")
    try:
        response = requests.get(KEV_URL)
        response.raise_for_status()
        kev_data = response.json()
        print("KEV list downloaded successfully.")
    except requests.exceptions.RequestException as e:
        print(f"Error downloading KEV list: {e}")
        return

    # --- 3. Identify new CVEs ---
    kev_vulnerabilities = kev_data.get("vulnerabilities", [])
    if existing_df is not None:
        existing_cves = set(existing_df['CVE ID'])
        new_vulnerabilities = [
            vuln for vuln in kev_vulnerabilities if vuln.get("cveID") not in existing_cves
        ]
        if not new_vulnerabilities:
            print("No new CVEs found in the KEV list. Exiting.")
            return
        print(f"Found {len(new_vulnerabilities)} new CVEs to add.")
        vulnerabilities_to_process = new_vulnerabilities
    else:
        print(f"Processing {len(kev_vulnerabilities)} CVEs from the KEV list.")
        vulnerabilities_to_process = kev_vulnerabilities

    # --- 4 & 5. Prepare data for DataFrame ---
    cve_rows = []
    for vuln in vulnerabilities_to_process:
        cve_rows.append({
            'CVE ID': vuln.get('cveID'),
            'Vulnerability Name': vuln.get('vulnerabilityName'),
            'CVSS Score': None,
            'CVSS Severity': None,
            'EPSS Score': None,  # Placeholder as EPSS is not in the KEV list
            'Vendor Project': vuln.get('vendorProject'),
            'Vendor Product': vuln.get('product'),
            'Date Added': vuln.get('dateAdded'),
            'Due Date': vuln.get('dueDate')
        })

    new_df = pd.DataFrame(cve_rows)

    # --- 6. Fetch NVD data in parallel ---
    print("Fetching CVSS data from NVD...")
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(fetch_nvd_data, cve_id): cve_id for cve_id in new_df['CVE ID']}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Fetching NVD Data"):
            cve_id = futures[future]
            try:
                score, severity = future.result()
                if score is not None:
                    new_df.loc[new_df['CVE ID'] == cve_id, 'CVSS Score'] = score
                    new_df.loc[new_df['CVE ID'] == cve_id, 'CVSS Severity'] = severity
            except Exception as e:
                print(f"Error processing {cve_id}: {e}")

    # --- Combine with existing data ---
    if existing_df is not None:
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
    else:
        combined_df = new_df

    # --- Create and format the spreadsheet ---
    print("Creating and formatting the spreadsheet...")
    writer = pd.ExcelWriter(SPREADSHEET_NAME, engine='openpyxl')
    combined_df.to_excel(writer, index=False, sheet_name='Enriched CVEs')
    workbook = writer.book
    worksheet = writer.sheets['Enriched CVEs']

    # --- 7 & 8. Apply color formatting ---
    for row_index, row in enumerate(combined_df.itertuples(), 2):
        # Color CVSS Score
        cvss_score_cell = worksheet.cell(row=row_index, column=combined_df.columns.get_loc('CVSS Score') + 1)
        score_color = get_color_for_score(getattr(row, 'CVSS_Score', None))
        if score_color:
            cvss_score_cell.fill = score_color

        # Color CVSS Severity
        cvss_severity_cell = worksheet.cell(row=row_index, column=combined_df.columns.get_loc('CVSS Severity') + 1)
        severity_color = get_color_for_severity(getattr(row, 'CVSS_Severity', None))
        if severity_color:
            cvss_severity_cell.fill = severity_color

    # --- 9. Add a timestamp ---
    timestamp_sheet = workbook.create_sheet("Metadata")
    timestamp_sheet['A1'] = "Last Updated"
    timestamp_sheet['B1'] = datetime.now().strftime(TIMESTAMP_FORMAT)

    writer.close()
    print(f"Successfully created/updated '{SPREADSHEET_NAME}'")

if __name__ == "__main__":
    main()